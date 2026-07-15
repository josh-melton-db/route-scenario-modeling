from __future__ import annotations

import io
import re
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from ..models import DeliveryDraft, DeliveryUploadError, DeliveryUploadResult

router = APIRouter(prefix="/scenarios/uploads", tags=["uploads"])

REQUIRED_COLUMNS = (
    "customer_name",
    "lat",
    "lng",
    "demand_cases",
    "service_minutes",
    "receiving_window_start",
    "receiving_window_end",
)
OPTIONAL_COLUMNS = ("delivery_day", "customer_id")
TIME_PATTERN = re.compile(r"^\d{1,2}:\d{2}$")


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_int(value: object, *, field: str, row_num: int) -> int:
    if value is None or (isinstance(value, float) and value != value):
        raise ValueError(f"Row {row_num}: {field} is required")
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_num}: {field} must be an integer") from exc


def _parse_float(value: object, *, field: str, row_num: int) -> float:
    if value is None or (isinstance(value, float) and value != value):
        raise ValueError(f"Row {row_num}: {field} is required")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_num}: {field} must be a number") from exc


def _parse_time(value: object, *, field: str, row_num: int, default: str) -> str:
    if value is None or value == "":
        return default
    text = str(value).strip()
    if not TIME_PATTERN.match(text):
        raise ValueError(f"Row {row_num}: {field} must look like HH:MM")
    hours, minutes = text.split(":")
    return f"{int(hours):02d}:{int(minutes):02d}"


def parse_deliveries_workbook(content: bytes) -> DeliveryUploadResult:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency declared in requirements
        raise HTTPException(status_code=500, detail="openpyxl is not installed") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return DeliveryUploadResult(deliveries=[], errors=[DeliveryUploadError(row=1, message="Workbook is empty")])

    headers = [_normalize_header(cell) for cell in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    missing = [col for col in REQUIRED_COLUMNS if col not in header_index]
    errors: list[DeliveryUploadError] = []
    if missing:
        return DeliveryUploadResult(
            deliveries=[],
            errors=[
                DeliveryUploadError(
                    row=1,
                    message=f"Missing required columns: {', '.join(missing)}",
                )
            ],
        )

    deliveries: list[DeliveryDraft] = []
    for excel_row_num, values in enumerate(rows[1:], start=2):
        if values is None or all(cell is None or str(cell).strip() == "" for cell in values):
            continue

        def cell(name: str) -> Any:
            idx = header_index.get(name)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        try:
            customer_name = str(cell("customer_name") or "").strip()
            if not customer_name:
                raise ValueError(f"Row {excel_row_num}: customer_name is required")
            lat = _parse_float(cell("lat"), field="lat", row_num=excel_row_num)
            lng = _parse_float(cell("lng"), field="lng", row_num=excel_row_num)
            if not (-90 <= lat <= 90):
                raise ValueError(f"Row {excel_row_num}: lat must be between -90 and 90")
            if not (-180 <= lng <= 180):
                raise ValueError(f"Row {excel_row_num}: lng must be between -180 and 180")
            demand_cases = _parse_int(cell("demand_cases"), field="demand_cases", row_num=excel_row_num)
            service_minutes = _parse_int(
                cell("service_minutes"),
                field="service_minutes",
                row_num=excel_row_num,
            )
            window_start = _parse_time(
                cell("receiving_window_start"),
                field="receiving_window_start",
                row_num=excel_row_num,
                default="08:00",
            )
            window_end = _parse_time(
                cell("receiving_window_end"),
                field="receiving_window_end",
                row_num=excel_row_num,
                default="16:00",
            )
            delivery_day_raw = cell("delivery_day")
            delivery_day = str(delivery_day_raw).strip() if delivery_day_raw not in (None, "") else None
            customer_id_raw = cell("customer_id")
            customer_id = str(customer_id_raw).strip() if customer_id_raw not in (None, "") else None
            deliveries.append(
                DeliveryDraft(
                    customer_name=customer_name,
                    lat=lat,
                    lng=lng,
                    demand_cases=demand_cases,
                    service_minutes=service_minutes,
                    receiving_window_start=window_start,
                    receiving_window_end=window_end,
                    delivery_day=delivery_day,
                    customer_id=customer_id,
                )
            )
        except ValueError as exc:
            errors.append(DeliveryUploadError(row=excel_row_num, message=str(exc)))

    return DeliveryUploadResult(deliveries=deliveries, errors=errors)


def build_template_bytes() -> bytes:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl is not installed") from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "deliveries"
    headers = list(REQUIRED_COLUMNS) + list(OPTIONAL_COLUMNS)
    sheet.append(headers)
    sheet.append(
        [
            "Acme Market",
            42.35,
            -83.05,
            90,
            30,
            "08:00",
            "16:00",
            "Tuesday",
            "",
        ]
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.post("/deliveries", response_model=DeliveryUploadResult)
async def upload_deliveries(file: UploadFile = File(...)) -> DeliveryUploadResult:
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an .xlsx Excel workbook.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    return parse_deliveries_workbook(content)


@router.get("/template")
async def download_template() -> StreamingResponse:
    content = build_template_bytes()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="delivery_upload_template.xlsx"'},
    )
